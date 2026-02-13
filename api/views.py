
import logging
import re
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.contrib.auth import authenticate
from django.db.models import Count, Q, F
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import joblib
import os
import sys
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import rsa, padding
from cryptography.hazmat.primitives import serialization
import base64
sys.path.append('..')
from ai_engine.train_model import load_model, predict_antibiotic
from users.models import User
from bacteria.models import Bacteria
from antibiotics.models import Antibiotic
from samples.models import Sample
from results.models import TestResult
from uploads.models import Upload
from ai_recommendations.models import AIRecommendation
from audit.models import AuditLog
from .serializers import (
    UserSerializer, BacteriaSerializer, AntibioticSerializer,
    SampleSerializer, TestResultSerializer, UploadSerializer,
    AIRecommendationSerializer, RegisterSerializer
)
from .permissions import AdminPermissions, DoctorPermissions, LabPermissions, ViewerPermissions, RoleBasedPermissions

# Set up logging
logger = logging.getLogger(__name__)

class CustomTokenObtainPairView(TokenObtainPairView):
    permission_classes = (AllowAny,)
    authentication_classes = []  # Disable authentication for login

    def post(self, request, *args, **kwargs):
        # Get credentials from request
        email = request.data.get('email') or request.data.get('username')
        password = request.data.get('password')

        # Authenticate user with email
        user = authenticate(request, username=email, password=password)

        if user is None:
            return Response(
                {"error": "Invalid credentials"},
                status=status.HTTP_401_UNAUTHORIZED
            )

        # Generate tokens
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)
        refresh_token = str(refresh)

        # Log login event
        ip_address = self.get_client_ip(request)
        AuditLog.objects.create(
            user_id=user,
            action_type='login',
            ip_address=ip_address
        )

        # Update last login
        user.last_login = timezone.now()
        user.save(update_fields=['last_login'])

        return Response({
            'refresh': refresh_token,
            'access': access_token,
            'user': UserSerializer(user).data
        })

    def get_client_ip(self, request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Logout user and log the event"""
        try:
            # Log logout event
            ip_address = self.get_client_ip(request)
            AuditLog.objects.create(
                user_id=request.user,
                action_type='logout',
                ip_address=ip_address
            )

            # For JWT, logout is handled client-side by discarding tokens
            # Optionally blacklist the token if using token blacklist
            return Response({"message": "Successfully logged out"})

        except Exception as e:
            logger = logging.getLogger(__name__)
            logger.error("Logout error", exc_info=True)
            return Response({"error": "Logout failed"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get_client_ip(self, request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

class RegisterView(APIView):
    permission_classes = (AllowAny,)

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({"message": "User registered successfully"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [RoleBasedPermissions]

class BacteriaViewSet(viewsets.ModelViewSet):
    queryset = Bacteria.objects.all()
    serializer_class = BacteriaSerializer
    permission_classes = [IsAuthenticated]

class AntibioticViewSet(viewsets.ModelViewSet):
    queryset = Antibiotic.objects.all()
    serializer_class = AntibioticSerializer
    permission_classes = [IsAuthenticated]

class SampleViewSet(viewsets.ModelViewSet):
    queryset = Sample.objects.all()
    serializer_class = SampleSerializer
    permission_classes = [IsAuthenticated]

class TestResultViewSet(viewsets.ModelViewSet):
    queryset = TestResult.objects.all()
    serializer_class = TestResultSerializer
    permission_classes = [IsAuthenticated]

class UploadViewSet(viewsets.ModelViewSet):
    queryset = Upload.objects.all()
    serializer_class = UploadSerializer
    permission_classes = [IsAuthenticated]

class AIRecommendationViewSet(viewsets.ModelViewSet):
    queryset = AIRecommendation.objects.all()
    serializer_class = AIRecommendationSerializer
    permission_classes = [IsAuthenticated]

class StatsView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            from core.data_service import create_data_service

            # Create data service with filters from request
            data_service = create_data_service(request)

            # Get statistics using unified service
            stats = data_service.get_statistics()

            # Add filter summary
            filter_summary = data_service.filter_engine.get_filter_summary()

            return Response({
                **stats,
                'filters': filter_summary
            })
        except Exception as e:
            logger.error(f"Error in StatsView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching statistics"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AnalyticsView(APIView):
    permission_classes = [DoctorPermissions]

    def get(self, request):
        # Placeholder for analytics data
        return Response({"message": "Analytics data"})

class ReportView(APIView):
    permission_classes = [DoctorPermissions]

    def get(self, request, report_type=None):
        from core.filters import create_filter_engine
        from core.data_service import UnifiedDataService

        # Create filter engine from request parameters
        filter_engine = create_filter_engine(request)
        data_service = UnifiedDataService(filter_engine)

        # Get filtered data using unified service
        filtered_results = data_service.get_filtered_results()
        stats = data_service.get_statistics()

        # Get filter summary for report title
        filter_summary = filter_engine.get_filter_summary()
        date_range = ""
        if filter_summary.get('active_filters', {}).get('date_from') and filter_summary.get('active_filters', {}).get('date_to'):
            date_range = f" ({filter_summary['active_filters']['date_from']} to {filter_summary['active_filters']['date_to']})"

        # Generate PDF report with charts using matplotlib
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            matplotlib.use('Agg')  # Use non-interactive backend
            from io import BytesIO
        except ImportError:
            # Fallback to text-only report if matplotlib not available
            response = HttpResponse(content_type='application/pdf')
            report_title = f"{report_type.title() if report_type else 'Antibiogram'} Report{date_range}"
            response['Content-Disposition'] = f'attachment; filename="{report_title.lower().replace(" ", "_")}_report.pdf"'

            p = canvas.Canvas(response)
            p.drawString(100, 750, report_title)
            p.drawString(100, 730, "Charts not available - matplotlib not installed")
            p.drawString(100, 710, "Please install matplotlib for visual charts")
            p.showPage()
            p.save()
            return response

        from reportlab.lib.pagesizes import letter
        from reportlab.lib.utils import ImageReader

        response = HttpResponse(content_type='application/pdf')
        report_title = f"{report_type.title() if report_type else 'Antibiogram'} Report{date_range}"
        response['Content-Disposition'] = f'attachment; filename="{report_title.lower().replace(" ", "_")}_report.pdf"'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter

        # Title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, height - 50, report_title)

        # Date range
        if date_range:
            p.setFont("Helvetica", 12)
            p.drawString(100, height - 70, f"Report Period:{date_range}")

        y_pos = height - 100

        # Basic stats
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y_pos, "Summary Statistics:")
        p.setFont("Helvetica", 12)
        y_pos -= 30

        p.drawString(120, y_pos, f"Total Samples: {stats['total_samples']}")
        y_pos -= 20
        p.drawString(120, y_pos, f"Total Bacteria: {stats['total_bacteria']}")
        y_pos -= 20
        p.drawString(120, y_pos, f"Total Antibiotics: {stats['total_antibiotics']}")
        y_pos -= 20
        p.drawString(120, y_pos, f"Filtered Results: {stats['total_results']}")
        y_pos -= 40

        # Sensitivity stats
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y_pos, "Sensitivity Distribution:")
        p.setFont("Helvetica", 12)
        y_pos -= 30
        p.drawString(120, y_pos, f"Resistant Cases: {stats['resistant_count']}")
        y_pos -= 20
        p.drawString(120, y_pos, f"Sensitive Cases: {stats['sensitive_count']}")
        y_pos -= 20
        p.drawString(120, y_pos, f"Intermediate Cases: {stats['intermediate_count']}")
        y_pos -= 50

        # Create charts
        chart_y_pos = y_pos

        # Chart 1: Sensitivity Distribution Pie Chart
        if y_pos < 400:
            p.showPage()
            y_pos = height - 100
            chart_y_pos = y_pos

        # Get sensitivity distribution from data service
        sensitivity_data = data_service.get_sensitivity_distribution()
        if sensitivity_data:
            fig, ax = plt.subplots(figsize=(6, 4))
            labels = [item['name'] for item in sensitivity_data]
            sizes = [item['value'] for item in sensitivity_data]
            colors = ['#ff9999','#66b3ff','#99ff99']  # Red for Resistant, Blue for Sensitive, Green for Intermediate
            ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
            ax.axis('equal')
            ax.set_title('Sensitivity Distribution')

            # Save chart to buffer
            buf = BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)

            # Add chart to PDF
            img = ImageReader(buf)
            p.drawImage(img, 100, chart_y_pos - 250, width=300, height=200)
            p.setFont("Helvetica-Bold", 12)
            p.drawString(100, chart_y_pos - 270, "Figure 1: Sensitivity Distribution")
            p.setFont("Helvetica", 10)
            p.drawString(100, chart_y_pos - 285, "This pie chart shows the proportion of test results by sensitivity category.")
            p.drawString(100, chart_y_pos - 300, "Color Legend: Red = Resistant, Blue = Sensitive, Green = Intermediate")
            p.drawString(100, chart_y_pos - 315, "Adjacent slices represent similar proportions; distant slices show significant differences.")
            chart_y_pos -= 350

        # Chart 2: Antibiotic Effectiveness Bar Chart
        if chart_y_pos < 300:
            p.showPage()
            chart_y_pos = height - 100

        # Get antibiotic effectiveness from data service
        effectiveness_data = data_service.get_antibiotic_effectiveness(top_n=10)
        if effectiveness_data:
            fig, ax = plt.subplots(figsize=(8, 4))
            antibiotics = [item['antibiotic'][:15] + '...' if len(item['antibiotic']) > 15 else item['antibiotic'] for item in effectiveness_data]
            effectiveness = [item['effectiveness'] for item in effectiveness_data]

            bars = ax.bar(range(len(antibiotics)), effectiveness, color='#4CAF50')
            ax.set_xlabel('Antibiotics')
            ax.set_ylabel('Effectiveness (%)')
            ax.set_title('Top 10 Antibiotic Effectiveness')
            ax.set_xticks(range(len(antibiotics)))
            ax.set_xticklabels(antibiotics, rotation=45, ha='right')

            # Add value labels on bars
            for bar, value in zip(bars, effectiveness):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                       f'{value}%', ha='center', va='bottom')

            # Save chart to buffer
            buf = BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)

            # Add chart to PDF
            img = ImageReader(buf)
            p.drawImage(img, 50, chart_y_pos - 300, width=500, height=250)
            p.setFont("Helvetica-Bold", 12)
            p.drawString(100, chart_y_pos - 320, "Figure 2: Antibiotic Effectiveness")
            p.setFont("Helvetica", 10)
            p.drawString(100, chart_y_pos - 335, "This bar chart shows the top 10 antibiotics ranked by effectiveness percentage.")
            p.drawString(100, chart_y_pos - 350, "Higher bars indicate better effectiveness; bars are sorted from most to least effective.")
            p.drawString(100, chart_y_pos - 365, "The green color represents effectiveness, with percentages shown above each bar.")
            chart_y_pos -= 400

        # Chart 3: Resistance Over Time Line Chart
        if chart_y_pos < 300:
            p.showPage()
            chart_y_pos = height - 100

        # Get resistance over time from data service
        resistance_trend = data_service.get_resistance_over_time(group_by='month')
        if resistance_trend:
            fig, ax = plt.subplots(figsize=(8, 4))
            dates = [item['month'] for item in resistance_trend]
            counts = [item['resistance'] for item in resistance_trend]

            ax.plot(range(len(dates)), counts, marker='o', linestyle='-', color='#f44336')
            ax.set_xlabel('Time Period')
            ax.set_ylabel('Resistance Cases')
            ax.set_title('Resistance Over Time')
            ax.set_xticks(range(len(dates)))
            ax.set_xticklabels(dates, rotation=45, ha='right')
            ax.grid(True, alpha=0.3)

            # Save chart to buffer
            buf = BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)

            # Add chart to PDF
            img = ImageReader(buf)
            p.drawImage(img, 50, chart_y_pos - 300, width=500, height=250)
            p.setFont("Helvetica-Bold", 12)
            p.drawString(100, chart_y_pos - 320, "Figure 3: Resistance Trends Over Time")
            p.setFont("Helvetica", 10)
            p.drawString(100, chart_y_pos - 335, "This line chart shows the number of resistant cases over time periods.")
            p.drawString(100, chart_y_pos - 350, "The red line connects data points chronologically; upward trends indicate increasing resistance.")
            p.drawString(100, chart_y_pos - 365, "Grid lines help track changes; markers show exact values for each time period.")

        p.showPage()
        p.save()
        return response

    def post(self, request):
        # Generate Excel report
        results = TestResult.objects.select_related('sample', 'sample__bacteria', 'antibiotic').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Antibiogram Report"

        # Header
        ws.cell(row=1, column=1, value="Sample ID")
        ws.cell(row=1, column=2, value="Bacteria")
        ws.cell(row=1, column=3, value="Antibiotic")
        ws.cell(row=1, column=4, value="Sensitivity")
        ws.cell(row=1, column=5, value="Date")

        # Data
        for row_num, result in enumerate(results, start=2):
            ws.cell(row=row_num, column=1, value=result.sample.id)
            ws.cell(row=row_num, column=2, value=result.sample.bacteria.name)
            ws.cell(row=row_num, column=3, value=result.antibiotic.name)
            ws.cell(row=row_num, column=4, value=result.sensitivity)
            ws.cell(row=row_num, column=5, value=result.sample.date.strftime('%Y-%m-%d'))

        # Save to response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="antibiogram_report.xlsx"'
        wb.save(response)
        return response

class SensitivityDistributionView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            from core.data_service import create_data_service

            # Create data service with filters from request
            data_service = create_data_service(request)

            # Get sensitivity distribution using unified service
            data = data_service.get_sensitivity_distribution()

            return Response(data)
        except Exception as e:
            logger.error(f"Error in SensitivityDistributionView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching sensitivity distribution"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AntibioticEffectivenessView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            from core.data_service import create_data_service

            # Create data service with filters from request
            data_service = create_data_service(request)

            # Get antibiotic effectiveness using unified service
            data = data_service.get_antibiotic_effectiveness(top_n=20)

            return Response(data)
        except Exception as e:
            logger.error(f"Error in AntibioticEffectivenessView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching antibiotic effectiveness"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ResistanceOverTimeView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            from core.data_service import create_data_service

            # Create data service with filters from request
            data_service = create_data_service(request)

            # Get resistance over time using unified service
            data = data_service.get_resistance_over_time(group_by='month')

            return Response(data)
        except Exception as e:
            logger.error(f"Error in ResistanceOverTimeView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching resistance over time data"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AIPredictView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        try:
            from core.data_service import create_data_service

            bacteria_name = request.data.get('bacteria_name')
            if not bacteria_name:
                return Response({"error": "Bacteria name is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Create data service with filters from request
            data_service = create_data_service(request)

            # Get antibiotic recommendations using unified service
            # This now reads ALL data from the database, not just a subset
            result = data_service.get_antibiotic_recommendations(bacteria_name, top_n=None)

            return Response(result)
        except Exception as e:
            logger.error(f"Error in AIPredictView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching AI predictions"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DigitalSignatureView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Create digital signature for document data"""
        try:
            document_data = request.data.get('document_data')
            if not document_data:
                return Response({"error": "Document data is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Generate RSA key pair
            private_key = rsa.generate_private_key(
                public_exponent=65537,
                key_size=2048,
            )
            public_key = private_key.public_key()

            # Sign the document
            signature = private_key.sign(
                document_data.encode('utf-8'),
                padding.PSS(
                    mgf=padding.MGF1(hashes.SHA256()),
                    salt_length=padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )

            # Serialize public key
            public_key_pem = public_key.public_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PublicFormat.SubjectPublicKeyInfo
            ).decode('utf-8')

            return Response({
                "signature": base64.b64encode(signature).decode('utf-8'),
                "public_key": public_key_pem,
                "algorithm": "RSA-PSS-SHA256"
            })

        except Exception as e:
            return Response({"error": f"Signature creation failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def put(self, request):
        """Verify digital signature"""
        try:
            document_data = request.data.get('document_data')
            signature_b64 = request.data.get('signature')
            public_key_pem = request.data.get('public_key')

            if not all([document_data, signature_b64, public_key_pem]):
                return Response({"error": "Document data, signature, and public key are required"}, status=status.HTTP_400_BAD_REQUEST)

            # Decode signature
            signature = base64.b64decode(signature_b64)

            # Load public key
            public_key = serialization.load_pem_public_key(public_key_pem.encode('utf-8'))

            # Verify signature
            try:
                public_key.verify(
                    signature,
                    document_data.encode('utf-8'),
                    padding.PSS(
                        mgf=padding.MGF1(hashes.SHA256()),
                        salt_length=padding.PSS.MAX_LENGTH
                    ),
                    hashes.SHA256()
                )
                return Response({"valid": True, "message": "Signature is valid"})
            except:
                return Response({"valid": False, "message": "Signature is invalid"})

        except Exception as e:
            return Response({"error": f"Signature verification failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BacteriaListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            bacteria = Bacteria.objects.all()
            data = [{'id': b.id, 'name': b.name, 'type': b.bacteria_type} for b in bacteria]
            return Response(data)
        except Exception as e:
            logger.error(f"Error in BacteriaListView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching bacteria list"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AntibioticListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            antibiotics = Antibiotic.objects.all()
            data = [{'id': a.id, 'name': a.name, 'category': a.category} for a in antibiotics]
            return Response(data)
        except Exception as e:
            logger.error(f"Error in AntibioticListView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching antibiotics list"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DepartmentListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            departments = Sample.objects.values_list('department', flat=True).distinct()
            data = [{'name': dept} for dept in departments if dept]
            return Response(data)
        except Exception as e:
            logger.error(f"Error in DepartmentListView: {str(e)}", exc_info=True)
            return Response({"error": "Internal server error occurred while fetching departments list"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ResistanceHeatmapView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            from core.data_service import create_data_service

            # Create data service with filters from request
            data_service = create_data_service(request)

            # Get resistance heatmap using unified service (with case-insensitive queries)
            heatmap_data = data_service.get_resistance_heatmap()

            return Response(heatmap_data)
        except Exception as e:
            logger.error(f"Error in ResistanceHeatmapView: {str(e)}", exc_info=True)
            return Response({"error": f"Internal server error occurred while fetching heatmap data: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class OCRView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Process image for OCR and extract antibiogram data"""
        try:
            import pytesseract
            from PIL import Image
            import numpy as np
            import re

            if 'image' not in request.FILES:
                return Response({"error": "No image file provided"}, status=status.HTTP_400_BAD_REQUEST)

            image_file = request.FILES['image']

            # Read image with PIL
            image = Image.open(image_file)

            # Convert to grayscale for better OCR
            gray_image = image.convert('L')

            # Convert PIL to numpy array for processing
            image_array = np.array(gray_image)

            # Apply simple thresholding for better contrast
            import cv2
            _, threshold_array = cv2.threshold(image_array, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

            # Convert back to PIL Image for pytesseract
            processed_image = Image.fromarray(threshold_array)

            # Extract text using Tesseract
            custom_config = r'--oem 3 --psm 6'
            text = pytesseract.image_to_string(processed_image, config=custom_config)

            # Parse antibiogram data
            parsed_data = self.parse_antibiogram_text(text)

            return Response({
                "text": text,
                "data": parsed_data
            })

        except ImportError as e:
            return Response({"error": f"OCR dependencies not installed: {str(e)}. Please install Tesseract OCR and ensure it's in your PATH."}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except pytesseract.pytesseract.TesseractNotFoundError:
            return Response({"error": "Tesseract OCR is not installed or not in PATH. Please install Tesseract OCR from https://github.com/UB-Mannheim/tesseract/wiki"}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        except Exception as e:
            return Response({"error": f"OCR processing failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def parse_antibiogram_text(self, text):
        """Parse extracted text to extract antibiogram data"""
        data = {}

        # Extract bacteria name
        bacteria_patterns = [
            r'Bacteria:\s*([^\n\r]+)',
            r'Organism:\s*([^\n\r]+)',
            r'Bacterium:\s*([^\n\r]+)',
            r'Pathogen:\s*([^\n\r]+)'
        ]

        for pattern in bacteria_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                data['bacteria'] = match.group(1).strip()
                break

        # Extract antibiotic and result pairs
        lines = text.split('\n')
        antibiotics = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Look for patterns like "Antibiotic: Result" or "Drug: Sensitivity"
            antibiotic_match = re.search(r'([A-Za-z\s]+(?:\s*\([^)]*\))?)[:\-]?\s*(Sensitive|Resistant|Intermediate|Susceptible)', line, re.IGNORECASE)
            if antibiotic_match:
                antibiotic_name = antibiotic_match.group(1).strip()
                sensitivity = antibiotic_match.group(2).strip()
                antibiotics.append({
                    'name': antibiotic_name,
                    'sensitivity': sensitivity
                })

        if antibiotics:
            data['antibiotics'] = antibiotics

        # Extract MIC values if present
        mic_pattern = r'MIC[:\-]?\s*([\d\.\<\>\=\sμg/mL]+)'
        mic_match = re.search(mic_pattern, text, re.IGNORECASE)
        if mic_match:
            data['mic'] = mic_match.group(1).strip()

        return data

class WelcomeView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        logger.info(f"Request received: {request.method} {request.path}")
        return Response({"message": "Welcome to the Data Analysis API Service!"})
